"""
Admin Reports API
==================
GET /api/admin/reports/fpo-summary/?format=excel   — download .xlsx
GET /api/admin/reports/fpo-summary/?format=pdf     — download .pdf

Filters:
  ?status=approved/submitted/rejected/draft/suspended
  ?district=TRS
  ?tier=A/B/C/D
  ?from_date=2026-01-01
  ?to_date=2026-06-30
"""

import io
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from django.template.loader import render_to_string

from drf_spectacular.utils import extend_schema, OpenApiParameter
from rest_framework import status
from rest_framework.views import APIView

from apps.core.utils.constants import UserRole, DISTRICTS_BILINGUAL
from apps.core.utils.responses import StandardResponse
from apps.database.models.fpo import FPO


def _can_view(user):
    return user.groups.filter(name__in=[UserRole.SUPER_ADMIN, UserRole.SUB_ADMIN]).exists()


def _build_queryset(params):
    qs = FPO.objects.filter(is_deleted=False).select_related('primary_user', 'primary_user__profile')

    status_val = params.get('status', '').strip()
    if status_val:
        qs = qs.filter(status=status_val)

    district = params.get('district', '').strip()
    if district:
        qs = qs.filter(district=district)

    tier = params.get('tier', '').strip().upper()
    if tier:
        qs = qs.filter(tier=tier)

    from_date = params.get('from_date', '').strip()
    if from_date:
        qs = qs.filter(created_at__date__gte=from_date)

    to_date = params.get('to_date', '').strip()
    if to_date:
        qs = qs.filter(created_at__date__lte=to_date)

    return qs.order_by('-created_at')


def _rows(qs):
    rows = []
    for fpo in qs:
        u = fpo.primary_user
        phone = ''
        if u:
            try:
                phone = u.profile.phone or ''
            except Exception:
                pass

        district_name = DISTRICTS_BILINGUAL.get(fpo.district, (fpo.district, fpo.district))[0] if fpo.district else ''

        rows.append({
            'application_id':  fpo.application_id or '—',
            'name':            fpo.name or '',
            'district':        district_name,
            'legal_structure': fpo.legal_structure or '',
            'status':          fpo.status or '',
            'tier':            fpo.tier or 'N/A',
            'total_members':   fpo.total_members or 0,
            'contact_name':    f'{u.first_name} {u.last_name}'.strip() if u else '',
            'contact_email':   u.email if u else '',
            'contact_phone':   phone,
            'registered_date': fpo.created_at.strftime('%d-%m-%Y') if fpo.created_at else '',
        })
    return rows


# ─────────────────────────────────────────────────────────────────────────────
# Excel
# ─────────────────────────────────────────────────────────────────────────────

_HEADERS = [
    'Application ID', 'FPO Name', 'District', 'Legal Structure',
    'Status', 'Tier', 'Total Members',
    'Contact Name', 'Contact Email', 'Contact Phone', 'Registered Date',
]

_HEADER_FILL  = PatternFill(start_color='2E7D32', end_color='2E7D32', fill_type='solid')
_HEADER_FONT  = Font(bold=True, color='FFFFFF', size=11)
_ALT_FILL     = PatternFill(start_color='F1F8E9', end_color='F1F8E9', fill_type='solid')


def _generate_excel(rows, title='FPO Summary Report'):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'FPO Summary'

    # Title row
    ws.merge_cells('A1:K1')
    title_cell = ws['A1']
    title_cell.value = title
    title_cell.font  = Font(bold=True, size=14, color='2E7D32')
    title_cell.alignment = Alignment(horizontal='center')
    ws.row_dimensions[1].height = 28

    # Sub-title
    ws.merge_cells('A2:K2')
    ws['A2'].value = f'Generated on {datetime.now().strftime("%d %B %Y, %I:%M %p")}'
    ws['A2'].font  = Font(italic=True, size=10, color='757575')
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.append([])  # blank row

    # Header row
    ws.append(_HEADERS)
    header_row = ws.max_row
    for col, cell in enumerate(ws[header_row], 1):
        cell.fill      = _HEADER_FILL
        cell.font      = _HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    ws.row_dimensions[header_row].height = 20

    # Data rows
    for i, row in enumerate(rows):
        ws.append([
            row['application_id'],
            row['name'],
            row['district'],
            row['legal_structure'],
            row['status'].replace('_', ' ').title(),
            row['tier'],
            row['total_members'],
            row['contact_name'],
            row['contact_email'],
            row['contact_phone'],
            row['registered_date'],
        ])
        if i % 2 == 0:
            for cell in ws[ws.max_row]:
                cell.fill = _ALT_FILL

    # Auto column widths
    col_widths = [18, 30, 18, 20, 14, 10, 14, 22, 28, 16, 16]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    # Freeze header
    ws.freeze_panes = f'A{header_row + 1}'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
# PDF (WeasyPrint)
# ─────────────────────────────────────────────────────────────────────────────

_PDF_HTML = """
<!DOCTYPE html>
<html>
<head>
<meta charset="utf-8">
<style>
  @page {{ size: A4 landscape; margin: 1cm; margin-left: 0.8cm; }}
  body {{ font-family: Arial, sans-serif; font-size: 9pt; color: #333; }}
  h1 {{ color: #2e7d32; font-size: 16pt; margin-bottom: 2px; }}
  .meta {{ color: #757575; font-size: 8pt; margin-bottom: 12px; }}
  table {{ width: 100%; border-collapse: collapse; table-layout: fixed; }}
  th {{ background: #2e7d32; color: #fff; padding: 5px 6px; text-align: left; font-size: 8pt; box-sizing: border-box;
        overflow: hidden; text-overflow: ellipsis; white-space: nowrap; }}
  td {{ padding: 4px 6px; border-bottom: 1px solid #e0e0e0; font-size: 8pt; box-sizing: border-box;
        word-break: break-word; overflow-wrap: break-word; }}
  th:nth-child(1)  {{ width: 10%; }}
  th:nth-child(2)  {{ width: 14%; }}
  th:nth-child(3)  {{ width: 8%;  }}
  th:nth-child(4)  {{ width: 10%; }}
  th:nth-child(5)  {{ width: 8%;  }}
  th:nth-child(6)  {{ width: 4%;  }}
  th:nth-child(7)  {{ width: 5%;  }}
  th:nth-child(8)  {{ width: 10%; }}
  th:nth-child(9)  {{ width: 13%; }}
  th:nth-child(10) {{ width: 8%;  }}
  th:nth-child(11) {{ width: 8%;  }}
  tr:nth-child(even) td {{ background: #f1f8e9; }}
  .badge-approved {{ color: #2e7d32; font-weight: bold; }}
  .badge-rejected {{ color: #c62828; font-weight: bold; }}
  .badge-suspended {{ color: #e65100; font-weight: bold; }}
  .badge-draft {{ color: #757575; }}
  .badge-submitted {{ color: #1565c0; font-weight: bold; }}
  .total {{ margin-top: 10px; font-size: 9pt; color: #555; }}
</style>
</head>
<body>
  <h1>FPO Summary Report</h1>
  <p class="meta">Generated on {generated_on} &nbsp;|&nbsp; Total records: {total}</p>
  <table>
    <thead>
      <tr>
        <th>Application ID</th>
        <th>FPO Name</th>
        <th>District</th>
        <th>Legal Structure</th>
        <th>Status</th>
        <th>Tier</th>
        <th>Members</th>
        <th>Contact Name</th>
        <th>Email</th>
        <th>Phone</th>
        <th>Registered</th>
      </tr>
    </thead>
    <tbody>
      {rows_html}
    </tbody>
  </table>
  <p class="total">Total: {total} FPOs</p>
</body>
</html>
"""

_STATUS_CLASSES = {
    'approved':     'badge-approved',
    'rejected':     'badge-rejected',
    'suspended':    'badge-suspended',
    'draft':        'badge-draft',
    'submitted':    'badge-submitted',
    'under_review': 'badge-submitted',
    'info_required':'badge-draft',
}


def _generate_pdf(rows):
    from weasyprint import HTML

    rows_html = ''
    for row in rows:
        css_class = _STATUS_CLASSES.get(row['status'], '')
        rows_html += (
            f'<tr>'
            f'<td>{row["application_id"]}</td>'
            f'<td>{row["name"]}</td>'
            f'<td>{row["district"]}</td>'
            f'<td>{row["legal_structure"].replace("_", " ").title()}</td>'
            f'<td class="{css_class}">{row["status"].replace("_", " ").title()}</td>'
            f'<td>{row["tier"]}</td>'
            f'<td>{row["total_members"]}</td>'
            f'<td>{row["contact_name"]}</td>'
            f'<td>{row["contact_email"]}</td>'
            f'<td>{row["contact_phone"]}</td>'
            f'<td>{row["registered_date"]}</td>'
            f'</tr>'
        )

    html_content = _PDF_HTML.format(
        generated_on=datetime.now().strftime('%d %B %Y, %I:%M %p'),
        total=len(rows),
        rows_html=rows_html,
    )

    buf = io.BytesIO()
    HTML(string=html_content).write_pdf(buf)
    buf.seek(0)
    return buf


# ─────────────────────────────────────────────────────────────────────────────
# View
# ─────────────────────────────────────────────────────────────────────────────

class FPOSummaryReportView(APIView):

    @extend_schema(
        tags=['Admin - Reports'],
        summary='Download FPO summary report',
        description=(
            'Downloads a full FPO summary report as Excel or PDF.\n\n'
            '**Format:** `?file_format=excel` (default) or `?file_format=pdf`\n\n'
            '**Filters:**\n'
            '- `?status=approved/submitted/rejected/draft/suspended`\n'
            '- `?district=TRS` (district code)\n'
            '- `?tier=A/B/C/D`\n'
            '- `?from_date=2026-01-01&to_date=2026-06-30`\n\n'
            'Returns a file download, not JSON.'
        ),
        parameters=[
            OpenApiParameter('file_format', description='excel or pdf (default: excel)', required=False, type=str),
            OpenApiParameter('status',    description='FPO status filter', required=False, type=str),
            OpenApiParameter('district',  description='District code', required=False, type=str),
            OpenApiParameter('tier',      description='Tier A/B/C/D', required=False, type=str),
            OpenApiParameter('from_date', description='YYYY-MM-DD', required=False, type=str),
            OpenApiParameter('to_date',   description='YYYY-MM-DD', required=False, type=str),
        ],
        responses={200: None},
    )
    def get(self, request):
        if not _can_view(request.user):
            return StandardResponse.error('Permission denied.', status_code=status.HTTP_403_FORBIDDEN)

        fmt = request.query_params.get('file_format', 'excel').lower().strip()
        if fmt not in ('excel', 'pdf'):
            return StandardResponse.error('file_format must be excel or pdf.',
                                          status_code=status.HTTP_400_BAD_REQUEST)

        qs   = _build_queryset(request.query_params)
        rows = _rows(qs)
        ts   = datetime.now().strftime('%Y%m%d_%H%M%S')

        if fmt == 'excel':
            buf      = _generate_excel(rows)
            filename = f'fpo_summary_{ts}.xlsx'
            response = HttpResponse(
                buf.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            return response

        buf      = _generate_pdf(rows)
        filename = f'fpo_summary_{ts}.pdf'
        response = HttpResponse(buf.read(), content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        return response
