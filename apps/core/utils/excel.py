"""
Excel and CSV utilities for KAU-FPO Platform
=============================================

Handles export (download template) and import (upload translations)
for both .xlsx and .csv formats.
"""

import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# =============================================================================
# EXPORT — Generate Template File
# =============================================================================

def generate_translation_template(keys_data, language_name, include_category=False, format='xlsx'):
    """
    Generate a translation template file for admin to fill.

    Args:
        keys_data: List of dicts with keys to translate
                   [{'category': 'auth', 'key': 'login_success',
                     'context': '...', 'english_value': '...'}]
        language_name: Display name of target language (e.g. 'Tamil')
        include_category: True if exporting all categories (adds category column)
        format: 'xlsx' or 'csv'

    Returns:
        BytesIO buffer with file content
    """
    if format == 'csv':
        return _generate_csv(keys_data, language_name, include_category)
    return _generate_xlsx(keys_data, language_name, include_category)


def _generate_xlsx(keys_data, language_name, include_category):
    """Generate Excel template with formatting."""
    wb = openpyxl.Workbook()

    # ── Sheet 1: Translations ──
    ws = wb.active
    ws.title = 'Translations'

    # Header styles
    header_font = Font(bold=True, color='FFFFFF')
    header_fill_blue = PatternFill(start_color='2563EB', end_color='2563EB', fill_type='solid')
    header_fill_green = PatternFill(start_color='16A34A', end_color='16A34A', fill_type='solid')
    locked_fill = PatternFill(start_color='F3F4F6', end_color='F3F4F6', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center')

    # Build headers
    headers = []
    if include_category:
        headers.append('category')
    headers += ['key', 'context', 'english_value', f'{language_name.lower()}_value']

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.alignment = center
        # Last column (translation) is green — admin fills this
        if col_idx == len(headers):
            cell.fill = header_fill_green
        else:
            cell.fill = header_fill_blue

    # Write data rows
    for row_idx, item in enumerate(keys_data, start=2):
        col = 1
        if include_category:
            cell = ws.cell(row=row_idx, column=col, value=item['category'])
            cell.fill = locked_fill
            col += 1

        # key
        cell = ws.cell(row=row_idx, column=col, value=item['key'])
        cell.fill = locked_fill
        col += 1

        # context
        cell = ws.cell(row=row_idx, column=col, value=item.get('context', ''))
        cell.fill = locked_fill
        col += 1

        # english_value
        cell = ws.cell(row=row_idx, column=col, value=item.get('english_value', ''))
        cell.fill = locked_fill
        col += 1

        # translation value — empty, admin fills this
        ws.cell(row=row_idx, column=col, value='')

    # Column widths
    column_widths = []
    if include_category:
        column_widths.append(15)
    column_widths += [25, 40, 45, 45]

    for col_idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 25

    # ── Sheet 2: Instructions ──
    ws2 = wb.create_sheet(title='Instructions')
    instructions = [
        ('HOW TO FILL THIS FILE', True),
        ('', False),
        ('1. Fill ONLY the last column with your translations', False),
        ('2. Do NOT change category, key, context, english_value columns', False),
        ('3. Do NOT add or delete rows', False),
        ('4. Leave the translation column empty to skip that key', False),
        ('5. Save as .xlsx or .csv before uploading', False),
        ('', False),
        ('COLUMN MEANINGS', True),
        ('', False),
        ('category       → Group the message belongs to (do not change)', False),
        ('key            → Unique identifier (do not change)', False),
        ('context        → Where this message appears (for reference only)', False),
        ('english_value  → English text (for reference only)', False),
        (f'{language_name.lower()}_value  → TYPE YOUR TRANSLATION HERE', False),
    ]

    for row_idx, (text, is_heading) in enumerate(instructions, start=1):
        cell = ws2.cell(row=row_idx, column=1, value=text)
        if is_heading:
            cell.font = Font(bold=True, size=12)

    ws2.column_dimensions['A'].width = 70

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _generate_csv(keys_data, language_name, include_category):
    """Generate CSV template."""
    buffer = io.StringIO()
    writer = csv.writer(buffer)

    # Headers
    headers = []
    if include_category:
        headers.append('category')
    headers += ['key', 'context', 'english_value', f'{language_name.lower()}_value']
    writer.writerow(headers)

    # Data rows
    for item in keys_data:
        row = []
        if include_category:
            row.append(item['category'])
        row += [
            item['key'],
            item.get('context', ''),
            item.get('english_value', ''),
            '',  # empty — admin fills
        ]
        writer.writerow(row)

    buffer.seek(0)
    return io.BytesIO(buffer.getvalue().encode('utf-8-sig'))  # utf-8-sig for Excel compatibility


# =============================================================================
# IMPORT — Parse Uploaded File
# =============================================================================

def parse_translation_file(file, language_code, category_code=None):
    """
    Parse uploaded Excel or CSV file.

    Args:
        file: Uploaded file object
        language_code: Target language code (e.g. 'ta')
        category_code: Optional category filter

    Returns:
        List of dicts: [{'category': str, 'key': str, 'value': str}]
        or raises ValueError with descriptive message
    """
    filename = file.name.lower()

    if filename.endswith('.xlsx') or filename.endswith('.xls'):
        return _parse_xlsx(file, language_code, category_code)
    elif filename.endswith('.csv'):
        return _parse_csv(file, language_code, category_code)
    else:
        raise ValueError("Unsupported file format. Please upload .xlsx or .csv file.")


def _parse_xlsx(file, language_code, category_code):
    """Parse Excel file."""
    try:
        wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
        ws = wb.active
    except Exception:
        raise ValueError("Invalid Excel file. Please download a fresh template and try again.")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise ValueError("File is empty.")

    return _parse_rows(rows, category_code)


def _parse_csv(file, language_code, category_code):
    """Parse CSV file."""
    try:
        content = file.read().decode('utf-8-sig')
        reader = csv.reader(io.StringIO(content))
        rows = list(reader)
    except Exception:
        raise ValueError("Invalid CSV file. Please check the file encoding and try again.")

    if not rows:
        raise ValueError("File is empty.")

    return _parse_rows(rows, category_code)


def _parse_rows(rows, category_code):
    """
    Parse rows from Excel or CSV into list of translation dicts.

    Expected columns (order matters):
    - With category:    category, key, context, english_value, translation_value
    - Without category: key, context, english_value, translation_value
    """
    if not rows:
        raise ValueError("File has no data.")

    headers = [str(h).strip().lower() if h else '' for h in rows[0]]

    # Detect if category column is present
    has_category = 'category' in headers

    # Find column indexes
    try:
        if has_category:
            cat_idx = headers.index('category')
            key_idx = headers.index('key')
        else:
            cat_idx = None
            key_idx = headers.index('key')

        # Last column is always the translation value
        val_idx = len(headers) - 1

    except ValueError:
        raise ValueError(
            "Invalid file format. Required columns: key, english_value, and translation column. "
            "Please download a fresh template."
        )

    results = []
    errors = []

    for row_num, row in enumerate(rows[1:], start=2):
        if not row or all(cell is None or str(cell).strip() == '' for cell in row):
            continue  # skip empty rows

        # Get values safely
        key = str(row[key_idx]).strip() if row[key_idx] else ''
        value = str(row[val_idx]).strip() if len(row) > val_idx and row[val_idx] else ''
        category = str(row[cat_idx]).strip() if cat_idx is not None and row[cat_idx] else category_code

        # Skip if key is empty
        if not key:
            errors.append({'row': row_num, 'reason': 'Empty key — row skipped'})
            continue

        # Skip if translation value is empty
        if not value:
            continue  # silently skip — admin left it blank intentionally

        # Override category if provided in request
        if category_code:
            category = category_code

        if not category:
            errors.append({'row': row_num, 'key': key, 'reason': 'No category found for this row'})
            continue

        results.append({
            'category': category,
            'key': key,
            'value': value,
        })

    return results, errors
