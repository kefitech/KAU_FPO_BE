"""
FPO Team Management API
========================
Manage secondary users within an FPO.

GET   /api/fpo/me/team/                      — list all team members
POST  /api/fpo/me/team/invite/               — invite single user
POST  /api/fpo/me/team/bulk-invite/          — invite multiple via JSON
POST  /api/fpo/me/team/bulk-invite-file/     — invite multiple via Excel/CSV
POST  /api/fpo/me/team/bulk-activate/        — activate multiple by user_ids
POST  /api/fpo/me/team/bulk-deactivate/      — deactivate multiple by user_ids
POST  /api/fpo/me/team/{id}/deactivate/      — deactivate single user

Rules:
- FPO must be APPROVED before inviting members
- Only the primary user can invite / activate / deactivate
- No secondary user limit (KAU confirmed)
- Invited users are auto-activated (no approval step)
- Invited users must change password on first login
"""

import csv
import io
import secrets

from django.contrib.auth import get_user_model
from django.contrib.auth.models import Group
from django.db import transaction
from django.http import HttpResponse

from drf_spectacular.utils import extend_schema
from rest_framework import serializers, status
from rest_framework.parsers import MultiPartParser
from rest_framework.views import APIView

import openpyxl

from apps.core.models.generic import AuditLog
from apps.core.permissions.rbac import IsFPOManager
from apps.core.services.audit import AuditService as AuditLogService
from apps.core.utils.constants import FPOStatus, UserRole
from apps.core.utils.responses import StandardResponse
from apps.database.models.fpo import FPO, FPOUserMembership
from apps.notifications.services import send_notification

User = get_user_model()


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _get_primary_fpo(user):
    """Return FPO only if this user is the primary user."""
    return FPO.objects.filter(primary_user=user, is_deleted=False).first()


# ─────────────────────────────────────────────────────────────────────────────
# Serializers
# ─────────────────────────────────────────────────────────────────────────────

class TeamInviteSerializer(serializers.Serializer):
    first_name = serializers.CharField(max_length=150)
    last_name  = serializers.CharField(max_length=150)
    email      = serializers.EmailField()
    phone      = serializers.CharField(max_length=10, required=False, allow_blank=True)

    def validate_email(self, value):
        if User.objects.filter(email__iexact=value).exists():
            raise serializers.ValidationError('A user with this email already exists.')
        return value.lower()


class TeamMemberSerializer(serializers.ModelSerializer):
    id         = serializers.IntegerField(source='user.id', read_only=True)
    first_name = serializers.CharField(source='user.first_name', read_only=True)
    last_name  = serializers.CharField(source='user.last_name', read_only=True)
    email      = serializers.EmailField(source='user.email', read_only=True)
    phone      = serializers.SerializerMethodField()
    role       = serializers.CharField(source='role.name', read_only=True)
    joined_at  = serializers.DateTimeField(read_only=True)

    class Meta:
        model  = FPOUserMembership
        fields = ['id', 'first_name', 'last_name', 'email', 'phone', 'role', 'is_active', 'joined_at']

    def get_phone(self, obj):
        profile = getattr(obj.user, 'profile', None)
        return profile.phone if profile else ''


# ─────────────────────────────────────────────────────────────────────────────
# Views
# ─────────────────────────────────────────────────────────────────────────────

class TeamListView(APIView):
    permission_classes = [IsFPOManager]

    @extend_schema(
        tags=['FPO - Team'],
        summary='List team members',
        description='Returns all secondary users in this FPO. Only accessible by the primary user.',
        responses={200: TeamMemberSerializer(many=True)},
    )
    def get(self, request):
        fpo = _get_primary_fpo(request.user)
        if not fpo:
            return StandardResponse.error(
                'Only the primary user can view the team.',
                status_code=status.HTTP_403_FORBIDDEN,
            )

        memberships = FPOUserMembership.objects.filter(
            fpo=fpo, is_deleted=False
        ).select_related('user', 'user__profile', 'role').exclude(
            user=request.user
        )
        serializer = TeamMemberSerializer(memberships, many=True)
        return StandardResponse.success(serializer.data, 'Team members retrieved.')


class TeamInviteView(APIView):
    permission_classes = [IsFPOManager]

    @extend_schema(
        tags=['FPO - Team'],
        summary='Invite secondary user',
        description=(
            'Creates a secondary user account and adds them to the FPO team. '
            'The invited user receives a welcome email with a temporary password '
            'and must change it on first login.\n\n'
            '**Rules:**\n'
            '- FPO must be in APPROVED status\n'
            '- Only the primary user can invite members\n'
            '- No limit on secondary users\n'
            '- Email must not already be registered'
        ),
        request=TeamInviteSerializer,
        responses={201: None},
    )
    def post(self, request):
        fpo = _get_primary_fpo(request.user)
        if not fpo:
            return StandardResponse.error(
                'Only the primary user can invite team members.',
                status_code=status.HTTP_403_FORBIDDEN,
            )

        if fpo.status != FPOStatus.APPROVED:
            return StandardResponse.error(
                'Team members can only be invited after the FPO is approved.',
                status_code=status.HTTP_400_BAD_REQUEST,
            )

        serializer = TeamInviteSerializer(data=request.data)
        if not serializer.is_valid():
            return StandardResponse.validation_error(errors=serializer.errors)

        data = serializer.validated_data
        temp_password = secrets.token_urlsafe(10)

        with transaction.atomic():
            # Create user account
            user = User.objects.create_user(
                username=data['email'],
                email=data['email'],
                first_name=data['first_name'],
                last_name=data['last_name'],
                password=temp_password,
            )

            # Assign fpo_manager + secondary group
            fpo_manager_group = Group.objects.get(name=UserRole.FPO_MANAGER)
            secondary_group   = Group.objects.get(name='secondary')
            user.groups.add(fpo_manager_group, secondary_group)

            # Set profile
            profile = getattr(user, 'profile', None)
            if profile:
                if data.get('phone'):
                    profile.phone = data['phone']
                profile.must_change_password = True
                profile.save(update_fields=['phone', 'must_change_password'])

            # Create membership — auto-activated
            membership = FPOUserMembership.objects.create(
                fpo=fpo,
                user=user,
                role=secondary_group,
                is_active=True,
                created_by=request.user,
            )

        # Send welcome notification
        lang = getattr(request, 'language', 'en')
        send_notification(
            user=user,
            code='welcome',
            channel='email',
            context={
                'user_name':    f'{data["first_name"]} {data["last_name"]}',
                'email':        data['email'],
                'temp_password': temp_password,
            },
            lang=lang,
        )

        # Audit log
        AuditLogService.log(
            user=request.user,
            action=AuditLog.Action.FPO_USER_INVITE,
            instance=fpo,
            request=request,
            changes={
                'invited_user':  data['email'],
                'invited_name':  f'{data["first_name"]} {data["last_name"]}',
            },
        )

        return StandardResponse.created(
            data={
                'id':         user.id,
                'email':      user.email,
                'first_name': user.first_name,
                'last_name':  user.last_name,
            },
            message='Team member invited successfully. A welcome email with login credentials has been sent.',
        )


class TeamDeactivateView(APIView):
    permission_classes = [IsFPOManager]

    @extend_schema(
        tags=['FPO - Team'],
        summary='Deactivate team member',
        description='Deactivates a secondary user. They will no longer be able to log in. Only the primary user can do this.',
        responses={200: None},
    )
    def post(self, request, user_id):
        fpo = _get_primary_fpo(request.user)
        if not fpo:
            return StandardResponse.error(
                'Only the primary user can deactivate team members.',
                status_code=status.HTTP_403_FORBIDDEN,
            )

        try:
            membership = FPOUserMembership.objects.select_related('user').get(
                fpo=fpo, user_id=user_id, is_deleted=False,
            )
        except FPOUserMembership.DoesNotExist:
            return StandardResponse.error(
                'Team member not found.',
                status_code=status.HTTP_404_NOT_FOUND,
            )

        if membership.user == request.user:
            return StandardResponse.error(
                'You cannot deactivate yourself.',
                status_code=status.HTTP_400_BAD_REQUEST,
            )

        if not membership.is_active:
            return StandardResponse.error(
                'This team member is already inactive.',
                status_code=status.HTTP_400_BAD_REQUEST,
            )

        membership.is_active = False
        membership.updated_by = request.user
        membership.save(update_fields=['is_active', 'updated_by'])

        # Also deactivate their Django account
        membership.user.is_active = False
        membership.user.save(update_fields=['is_active'])

        AuditLogService.log(
            user=request.user,
            action=AuditLog.Action.FPO_USER_DEACTIVATE,
            instance=fpo,
            request=request,
            changes={'deactivated_user': membership.user.email},
        )

        return StandardResponse.success(
            message=f'{membership.user.get_full_name()} has been deactivated.',
        )


class TeamResetPasswordView(APIView):
    permission_classes = [IsFPOManager]

    @extend_schema(
        tags=['FPO - Team'],
        summary='Reset secondary user password',
        description=(
            'Generates a new temporary password for a secondary user and sends it to their email. '
            'The primary user never sees the password — it goes directly to the secondary user. '
            'The secondary user is forced to change it on next login.'
        ),
        responses={200: None},
    )
    def post(self, request, user_id):
        fpo = _get_primary_fpo(request.user)
        if not fpo:
            return StandardResponse.error(
                'Only the primary user can reset team member passwords.',
                status_code=status.HTTP_403_FORBIDDEN,
            )

        try:
            membership = FPOUserMembership.objects.select_related('user', 'user__profile').get(
                fpo=fpo, user_id=user_id, is_deleted=False,
            )
        except FPOUserMembership.DoesNotExist:
            return StandardResponse.error('Team member not found.', status_code=status.HTTP_404_NOT_FOUND)

        if membership.user == request.user:
            return StandardResponse.error('You cannot reset your own password here.',
                                          status_code=status.HTTP_400_BAD_REQUEST)

        temp_password = secrets.token_urlsafe(10)
        membership.user.set_password(temp_password)
        membership.user.save(update_fields=['password'])

        profile = getattr(membership.user, 'profile', None)
        if profile:
            profile.must_change_password = True
            profile.save(update_fields=['must_change_password'])

        lang = getattr(request, 'language', 'en')
        send_notification(
            user=membership.user,
            code='welcome',
            channel='email',
            context={
                'user_name':     membership.user.get_full_name(),
                'email':         membership.user.email,
                'temp_password': temp_password,
            },
            lang=lang,
        )

        return StandardResponse.success(
            message=f'Password reset. New credentials have been sent to {membership.user.email}.',
        )


# ─────────────────────────────────────────────────────────────────────────────
# Shared invite helper
# ─────────────────────────────────────────────────────────────────────────────

def _create_member(fpo, row, inviter, lang):
    """
    Create one secondary user from a dict with keys:
    first_name, last_name, email, phone (optional).

    Returns (user, temp_password) on success.
    Raises ValueError with a human-readable message on failure.
    """
    email = row.get('email', '').strip().lower()
    first_name = row.get('first_name', '').strip()
    last_name  = row.get('last_name', '').strip()
    phone      = row.get('phone', '').strip()

    if not email or not first_name or not last_name:
        raise ValueError('first_name, last_name and email are required.')

    if User.objects.filter(email__iexact=email).exists():
        raise ValueError(f'{email} is already registered.')

    temp_password = secrets.token_urlsafe(10)

    with transaction.atomic():
        user = User.objects.create_user(
            username=email,
            email=email,
            first_name=first_name,
            last_name=last_name,
            password=temp_password,
        )
        fpo_manager_group = Group.objects.get(name=UserRole.FPO_MANAGER)
        secondary_group   = Group.objects.get(name='secondary')
        user.groups.add(fpo_manager_group, secondary_group)

        profile = getattr(user, 'profile', None)
        if profile:
            if phone:
                profile.phone = phone
            profile.must_change_password = True
            profile.save(update_fields=['phone', 'must_change_password'])

        FPOUserMembership.objects.create(
            fpo=fpo,
            user=user,
            role=secondary_group,
            is_active=True,
            created_by=inviter,
        )

    send_notification(
        user=user,
        code='welcome',
        channel='email',
        context={
            'user_name':     f'{first_name} {last_name}',
            'email':         email,
            'temp_password': temp_password,
        },
        lang=lang,
    )

    return user, temp_password


# ─────────────────────────────────────────────────────────────────────────────
# Bulk invite — JSON
# ─────────────────────────────────────────────────────────────────────────────

class BulkInviteSerializer(serializers.Serializer):
    members = serializers.ListField(
        child=serializers.DictField(),
        min_length=1,
        max_length=100,
        help_text='List of members. Each must have first_name, last_name, email. phone is optional.',
    )


class TeamBulkInviteView(APIView):
    permission_classes = [IsFPOManager]

    @extend_schema(
        tags=['FPO - Team'],
        summary='Bulk invite team members (JSON)',
        description=(
            'Invite multiple secondary users in one call. '
            'Each row is processed independently — failures do not block others.\n\n'
            '**Request body:**\n'
            '```json\n'
            '{ "members": [\n'
            '  { "first_name": "Rajan", "last_name": "Kumar", "email": "rajan@example.com", "phone": "9876543210" },\n'
            '  { "first_name": "Meera", "last_name": "S", "email": "meera@example.com" }\n'
            '] }\n'
            '```\n\n'
            'Returns a summary with per-row success/failure details.'
        ),
        request=BulkInviteSerializer,
        responses={200: None},
    )
    def post(self, request):
        fpo = _get_primary_fpo(request.user)
        if not fpo:
            return StandardResponse.error(
                'Only the primary user can invite team members.',
                status_code=status.HTTP_403_FORBIDDEN,
            )
        if fpo.status != FPOStatus.APPROVED:
            return StandardResponse.error(
                'Team members can only be invited after the FPO is approved.',
                status_code=status.HTTP_400_BAD_REQUEST,
            )

        serializer = BulkInviteSerializer(data=request.data)
        if not serializer.is_valid():
            return StandardResponse.validation_error(errors=serializer.errors)

        lang    = getattr(request, 'language', 'en')
        members = serializer.validated_data['members']
        success, failed = [], []

        for i, row in enumerate(members, start=1):
            try:
                user, _ = _create_member(fpo, row, request.user, lang)
                success.append({'row': i, 'email': user.email, 'name': user.get_full_name()})
            except Exception as e:
                failed.append({'row': i, 'email': row.get('email', ''), 'reason': str(e)})

        if success:
            AuditLogService.log(
                user=request.user,
                action=AuditLog.Action.FPO_USER_INVITE,
                instance=fpo,
                request=request,
                changes={'bulk_invited': [r['email'] for r in success]},
            )

        return StandardResponse.success(
            data={'success': len(success), 'failed': len(failed), 'results': success, 'errors': failed},
            message=f'{len(success)} member(s) invited successfully, {len(failed)} failed.',
        )


# ─────────────────────────────────────────────────────────────────────────────
# Bulk invite — Excel / CSV file
# ─────────────────────────────────────────────────────────────────────────────

class TeamBulkInviteFileView(APIView):
    permission_classes = [IsFPOManager]
    parser_classes     = [MultiPartParser]

    @extend_schema(
        tags=['FPO - Team'],
        summary='Bulk invite team members (Excel/CSV)',
        description=(
            'Upload an `.xlsx` or `.csv` file to invite multiple secondary users.\n\n'
            '**Required columns:** `first_name`, `last_name`, `email`\n'
            '**Optional column:** `phone`\n\n'
            'Row 1 must be the header row. Each data row is processed independently.\n'
            'Returns a summary with per-row success/failure details.'
        ),
        responses={200: None},
    )
    def post(self, request):
        fpo = _get_primary_fpo(request.user)
        if not fpo:
            return StandardResponse.error(
                'Only the primary user can invite team members.',
                status_code=status.HTTP_403_FORBIDDEN,
            )
        if fpo.status != FPOStatus.APPROVED:
            return StandardResponse.error(
                'Team members can only be invited after the FPO is approved.',
                status_code=status.HTTP_400_BAD_REQUEST,
            )

        file = request.FILES.get('file')
        if not file:
            return StandardResponse.error('No file uploaded. Send file as multipart form field "file".',
                                          status_code=status.HTTP_400_BAD_REQUEST)

        filename = file.name.lower()
        rows = []

        try:
            if filename.endswith('.csv'):
                content = file.read().decode('utf-8-sig')
                reader  = csv.DictReader(io.StringIO(content))
                rows    = list(reader)
            elif filename.endswith('.xlsx'):
                wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
                ws = wb.active
                headers = [str(c.value).strip().lower() if c.value else '' for c in next(ws.iter_rows())]
                for row in ws.iter_rows(min_row=2, values_only=True):
                    rows.append(dict(zip(headers, [str(v).strip() if v is not None else '' for v in row])))
            else:
                return StandardResponse.error('Only .xlsx and .csv files are supported.',
                                              status_code=status.HTTP_400_BAD_REQUEST)
        except Exception as e:
            return StandardResponse.error(f'Could not parse file: {e}',
                                          status_code=status.HTTP_400_BAD_REQUEST)

        if not rows:
            return StandardResponse.error('File is empty or has no data rows.',
                                          status_code=status.HTTP_400_BAD_REQUEST)

        lang = getattr(request, 'language', 'en')
        success, failed = [], []

        for i, row in enumerate(rows, start=2):  # start=2 because row 1 is header
            try:
                user, _ = _create_member(fpo, row, request.user, lang)
                success.append({'row': i, 'email': user.email, 'name': user.get_full_name()})
            except Exception as e:
                failed.append({'row': i, 'email': row.get('email', ''), 'reason': str(e)})

        if success:
            AuditLogService.log(
                user=request.user,
                action=AuditLog.Action.FPO_USER_INVITE,
                instance=fpo,
                request=request,
                changes={'bulk_invited_via_file': [r['email'] for r in success]},
            )

        return StandardResponse.success(
            data={'success': len(success), 'failed': len(failed), 'results': success, 'errors': failed},
            message=f'{len(success)} member(s) invited successfully, {len(failed)} failed.',
        )


# ─────────────────────────────────────────────────────────────────────────────
# Bulk activate / deactivate
# ─────────────────────────────────────────────────────────────────────────────

class BulkActionSerializer(serializers.Serializer):
    user_ids = serializers.ListField(
        child=serializers.IntegerField(),
        min_length=1,
        help_text='List of user IDs to activate or deactivate.',
    )


class TeamBulkActivateView(APIView):
    permission_classes = [IsFPOManager]

    @extend_schema(
        tags=['FPO - Team'],
        summary='Bulk activate team members',
        description='Activate multiple secondary users by their user IDs.',
        request=BulkActionSerializer,
        responses={200: None},
    )
    def post(self, request):
        return _bulk_toggle(request, activate=True)


class TeamBulkDeactivateView(APIView):
    permission_classes = [IsFPOManager]

    @extend_schema(
        tags=['FPO - Team'],
        summary='Bulk deactivate team members',
        description='Deactivate multiple secondary users by their user IDs.',
        request=BulkActionSerializer,
        responses={200: None},
    )
    def post(self, request):
        return _bulk_toggle(request, activate=False)


def _bulk_toggle(request, activate: bool):
    fpo = _get_primary_fpo(request.user)
    if not fpo:
        return StandardResponse.error(
            'Only the primary user can manage team members.',
            status_code=status.HTTP_403_FORBIDDEN,
        )

    serializer = BulkActionSerializer(data=request.data)
    if not serializer.is_valid():
        return StandardResponse.validation_error(errors=serializer.errors)

    user_ids = serializer.validated_data['user_ids']
    success, failed = [], []

    for uid in user_ids:
        if uid == request.user.id:
            failed.append({'user_id': uid, 'reason': 'Cannot modify yourself.'})
            continue
        try:
            membership = FPOUserMembership.objects.select_related('user').get(
                fpo=fpo, user_id=uid, is_deleted=False,
            )
            membership.is_active      = activate
            membership.updated_by     = request.user
            membership.user.is_active = activate
            membership.save(update_fields=['is_active', 'updated_by'])
            membership.user.save(update_fields=['is_active'])
            success.append({'user_id': uid, 'name': membership.user.get_full_name()})
        except FPOUserMembership.DoesNotExist:
            failed.append({'user_id': uid, 'reason': 'Team member not found.'})

    if success:
        action = AuditLog.Action.FPO_USER_ACTIVATE if activate else AuditLog.Action.FPO_USER_DEACTIVATE
        AuditLogService.log(
            user=request.user,
            action=action,
            instance=fpo,
            request=request,
            changes={'user_ids': [r['user_id'] for r in success]},
        )

    verb = 'activated' if activate else 'deactivated'
    return StandardResponse.success(
        data={'success': len(success), 'failed': len(failed), 'results': success, 'errors': failed},
        message=f'{len(success)} member(s) {verb}, {len(failed)} failed.',
    )
